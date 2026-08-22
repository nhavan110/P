"""
Quét folder Google Drive "đc/" theo cấu trúc:

    đc/
      TCB/TCB.xlsx
      HPG/HPG.xlsx
      FPT/FPT.xlsx
      ...

Mỗi subfolder = 1 mã doanh nghiệp, chứa 1 file .xlsx với nhiều sheet
(balance_sheet, income_statement, cash_flow, chi_so_tai_chinh, ...).

Script tải từng file, đọc toàn bộ sheet bằng openpyxl (giá trị đã tính, không
công thức), rồi ghi ra data/companies.json để trang web (tab "Phân tích doanh
nghiệp") đọc tĩnh — không gọi Google Drive API trực tiếp từ trình duyệt.

Biến môi trường cần có:
  GOOGLE_SERVICE_ACCOUNT_JSON        (dùng chung với upload_to_drive.py)
  GOOGLE_DRIVE_COMPANIES_FOLDER_ID   ID của folder "đc/" (share quyền Viewer
                                      trở lên cho email service account)

Nếu thiếu GOOGLE_DRIVE_COMPANIES_FOLDER_ID, script sẽ bỏ qua (không lỗi) để
không chặn phần cập nhật portfolio chính khi tính năng này chưa được cấu hình.
"""

import io
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

from upload_to_drive import load_credentials_info  # tái dùng hàm parse secret

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
OUTPUT_PATH = Path(__file__).resolve().parent / "data" / "companies.json"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"
FOLDER_MIME = "application/vnd.google-apps.folder"


def list_children(drive, folder_id):
    """Liệt kê toàn bộ file/folder con trực tiếp của folder_id (có phân trang)."""
    items = []
    page_token = None
    while True:
        resp = drive.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            spaces="drive",
            fields="nextPageToken, files(id, name, mimeType)",
            pageSize=100,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


def download_xlsx_bytes(drive, file_meta):
    """Tải nội dung file xlsx (hoặc export nếu là Google Sheets) về bytes."""
    file_id = file_meta["id"]
    if file_meta["mimeType"] == GSHEET_MIME:
        request = drive.files().export_media(fileId=file_id, mimeType=XLSX_MIME)
    else:
        request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)

    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf.read()


def cell_to_json(value):
    """Chuẩn hoá giá trị cell về kiểu JSON-serializable, giữ nguyên number/text."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def fill_merged_cells(ws):
    """Điền giá trị ô gốc của các vùng merge vào toàn bộ vùng, để không bị
    mất dữ liệu khi đọc theo hàng/cột."""
    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        top_left = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                ws.cell(row=r, column=c).value = top_left


def read_workbook_sheets(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            continue  # bỏ qua sheet trống
        fill_merged_cells(ws)
        rows = [
            [cell_to_json(v) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
        # bỏ các hàng hoàn toàn trống ở cuối
        while rows and all(v is None for v in rows[-1]):
            rows.pop()
        sheets.append({"name": name, "rows": rows})
    return sheets


def main():
    service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    root_folder_id = os.environ.get("GOOGLE_DRIVE_COMPANIES_FOLDER_ID")

    if not root_folder_id:
        print(
            "ℹ️  Chưa cấu hình GOOGLE_DRIVE_COMPANIES_FOLDER_ID — bỏ qua bước "
            "cập nhật data/companies.json (tab Phân tích doanh nghiệp)."
        )
        return
    if not service_account_json:
        raise RuntimeError("Missing GOOGLE_SERVICE_ACCOUNT_JSON environment variable")

    credentials_info = load_credentials_info(service_account_json)
    print(f"ℹ️  Đăng nhập bằng service account: {credentials_info.get('client_email')}")

    credentials = service_account.Credentials.from_service_account_info(
        credentials_info,
        scopes=SCOPES,
    )
    drive = build("drive", "v3", credentials=credentials)

    try:
        top_level = list_children(drive, root_folder_id)
    except HttpError as e:
        print(f"❌ Lỗi khi đọc folder 'đc/' (folder_id={root_folder_id}):", file=sys.stderr)
        print(e.error_details if hasattr(e, "error_details") else e, file=sys.stderr)
        print(
            "→ Kiểm tra: (1) GOOGLE_DRIVE_COMPANIES_FOLDER_ID đúng chưa, "
            f"(2) folder đó đã share quyền Viewer cho {credentials_info.get('client_email')} chưa.",
            file=sys.stderr,
        )
        raise

    company_folders = [f for f in top_level if f["mimeType"] == FOLDER_MIME]
    if not company_folders:
        print("⚠️  Không tìm thấy subfolder doanh nghiệp nào trong 'đc/'.")

    companies = {}
    for folder in sorted(company_folders, key=lambda f: f["name"]):
        ticker = folder["name"].strip().upper()
        children = list_children(drive, folder["id"])
        xlsx_files = [
            f for f in children
            if f["mimeType"] in (XLSX_MIME, GSHEET_MIME) or f["name"].lower().endswith(".xlsx")
        ]
        if not xlsx_files:
            print(f"⚠️  {ticker}: không tìm thấy file .xlsx trong folder, bỏ qua.")
            continue

        # Ưu tiên file trùng tên với ticker (vd TCB/TCB.xlsx), nếu không lấy file đầu tiên.
        target = next(
            (f for f in xlsx_files if f["name"].lower() == f"{ticker.lower()}.xlsx"),
            xlsx_files[0],
        )

        try:
            xlsx_bytes = download_xlsx_bytes(drive, target)
            sheets = read_workbook_sheets(xlsx_bytes)
        except Exception as e:  # noqa: BLE001 - báo lỗi nhưng vẫn xử lý các mã khác
            print(f"❌ {ticker}: lỗi khi tải/đọc file '{target['name']}': {e}", file=sys.stderr)
            continue

        companies[ticker] = {"file_name": target["name"], "sheets": sheets}
        print(f"✅ {ticker}: đọc {len(sheets)} sheet từ '{target['name']}'")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "companies": companies,
    }
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"✅ Đã ghi {OUTPUT_PATH} ({len(companies)} doanh nghiệp)")


if __name__ == "__main__":
    main()
