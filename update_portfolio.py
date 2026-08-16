# ── IMPORTS ──────────────────────────────────────────────────────────────────
import json
from copy import copy
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import brentq, newton
from vnstock import Vnstock

# ── WORKAROUND: bug trong vnstock.core.utils.env.get_hosting_service() ──────
# Hàm này không có nhánh else, nên khi chạy trên môi trường không phải
# Colab/Codespace/Replit/Kaggle/HF Spaces (vd: GitHub Actions runner), biến
# `hosting_service` không được gán -> UnboundLocalError. Lỗi này bị tenacity
# retry rồi che thành `tenacity.RetryError` khi gọi stock.quote.history().
# Patch lại hàm để trả về giá trị mặc định an toàn thay vì crash.
import vnstock.core.utils.env as _vnstock_env

_original_get_hosting_service = _vnstock_env.get_hosting_service


def _safe_get_hosting_service():
    try:
        result = _original_get_hosting_service()
    except UnboundLocalError:
        result = None
    return result or "Local or Unknown"


_vnstock_env.get_hosting_service = _safe_get_hosting_service

# ── CẤU HÌNH ─────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).resolve().parent
PRICE_HISTORY_FILE     = BASE_DIR / "data" / "price_history.csv"      # giá thô (input/cache), KHÔNG chứa W/D hay kết quả
CASHFLOWS_FILE          = BASE_DIR / "data" / "cashflows.csv"          # sổ nạp/rút tiền (W/D), tích luỹ dần
PORTFOLIO_HISTORY_FILE  = BASE_DIR / "data" / "portfolio_history.csv"  # output: toàn bộ dữ liệu đã tính, ghi lại mỗi lần chạy
OUTPUT_FILE  = BASE_DIR / "data" / "MyPortfolio.xlsx"                  # output: bản Excel định dạng đẹp để xem/Drive
JSON_FILE    = BASE_DIR / "data" / "data.json"   # trong thư mục data/, khớp với fetch('data/data.json') trong index.html và được workflow "git add data/" commit
MANUAL_FILE  = BASE_DIR / "manual_entries.csv"       # inbox nhập W/D thủ công, tự xoá sau khi được gộp vào cashflows.csv
TRANSACTIONS_FILE = BASE_DIR / "data" / "transactions.csv"  # sổ giao dịch mua/bán CP + margin — SOURCE OF TRUTH của holdings
SHEET_NAME   = "Sheet1"
VNI_START    = "2022-06-01"
API_TEMPLATE = "https://api.simplize.vn/api/historical/quote/prices/{}?page=0&size=600"
DEFAULT_STOCK_CODES = ["HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB"]

RISK_FREE_RATE = 0.04
TRADING_DAYS   = 252

# Map cứng mã CP → ngành, dùng cho Allocation/Position Contribution (mục 4)
SECTOR_MAP = {
    "HPG": "Thép", "TCB": "Ngân hàng", "MBB": "Ngân hàng",
    "FPT": "Công nghệ", "PNJ": "Bán lẻ", "FRT": "Bán lẻ", "MWG": "Bán lẻ",
}


# ── HÀM TIỆN ÍCH DÙNG CHUNG ───────────────────────────────────────────────────
def safe_num(x):
    """Trả về float hợp lệ, hoặc None nếu là NaN/Infinity (JSON chuẩn không
    chấp nhận NaN/Infinity — trình duyệt sẽ báo lỗi parse nếu để lọt vào)."""
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(v):
        return None
    return v


# ── HÀM XIRR (mục 5) — dùng để tính lãi suất nội bộ trên chuỗi cashflow không
# đều theo thời gian. Đặt sớm vì không phụ thuộc df_pr, chỉ nhận list (date, amount).
def xirr(cashflows: list) -> float:
    """cashflows: list các (ngày, số tiền có dấu). Âm = tiền ra khỏi túi NĐT
    (D, hoặc E0 giả định mua vào đầu kỳ). Dương = tiền về túi NĐT
    (W, hoặc E1 giả định bán ra cuối kỳ)."""
    if len(cashflows) < 2:
        return None
    d0 = min(d for d, _ in cashflows)

    def npv(rate):
        return sum(cf / (1 + rate) ** ((d - d0).days / 365) for d, cf in cashflows)

    try:
        return newton(npv, 0.1, maxiter=100)
    except (RuntimeError, OverflowError):
        try:
            return brentq(npv, -0.9999, 10)
        except ValueError:
            return None


# ── 1. ĐỌC LỊCH SỬ GIÁ (price_history.csv — chỉ giá thô) ────────────────────
def load_price_history(path: Path) -> pd.DataFrame:
    cols = ["date", *DEFAULT_STOCK_CODES, "VNINDEX"]
    if not path.exists():
        return pd.DataFrame(columns=cols)
    df = pd.read_csv(path)
    df["date"] = pd.to_datetime(df["date"], format="%d/%m/%Y")
    return df.sort_values("date").reset_index(drop=True)


price_hist   = load_price_history(PRICE_HISTORY_FILE)
stock_codes  = [c for c in price_hist.columns if len(str(c)) == 3] or DEFAULT_STOCK_CODES


# ── 2. LẤY GIÁ CỔ PHIẾU TỪ API ───────────────────────────────────────────────
def fetch_stock_prices(codes: list) -> pd.DataFrame:
    """Lấy giá đóng cửa của danh sách mã CK, trả về DataFrame indexed by date."""
    frames = []
    for code in codes:
        resp = requests.get(API_TEMPLATE.format(code), timeout=10)
        resp.raise_for_status()
        df = pd.DataFrame(resp.json()["data"])
        df["date"] = pd.to_datetime(df["date"], unit="s")
        frames.append(df.set_index("date")["priceClose"].rename(code))
    return pd.concat(frames, axis=1).reset_index()   # outer join tự động


df_prices = fetch_stock_prices(stock_codes)


# ── 3. LẤY VNINDEX ────────────────────────────────────────────────────────────
today_str = date.today().strftime("%Y-%m-%d")
stock = Vnstock().stock(symbol="VNINDEX", source="VCI")
df_vni = (
    stock.quote.history(start=VNI_START, end=today_str)
    .rename(columns={"time": "date", "close": "VNINDEX"})
    .assign(date=lambda x: pd.to_datetime(x["date"]))
    .iloc[::-1]
    [["date", "VNINDEX"]]
)

df_prices["date"] = pd.to_datetime(df_prices["date"])
df_all = df_prices.merge(df_vni, on="date", how="left")


# ── 4. NỐI VỚI DỮ LIỆU CŨ, LƯU LẠI price_history.csv (chỉ giá thô) ──────────
last_update = price_hist["date"].max() if not price_hist.empty else (pd.Timestamp(VNI_START) - pd.Timedelta(days=1))

df_new     = df_all[df_all["date"] > last_update]
price_hist = (
    pd.concat([price_hist, df_new], ignore_index=True)
    .drop_duplicates(subset="date", keep="last")
    .sort_values("date")
    .fillna(0)
    .reset_index(drop=True)
)

price_hist_out = price_hist.copy()
price_hist_out["date"] = price_hist_out["date"].dt.strftime("%d/%m/%Y")
PRICE_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
price_hist_out.to_csv(PRICE_HISTORY_FILE, index=False)
print(f"✅ Đã lưu: {PRICE_HISTORY_FILE} ({len(price_hist_out)} dòng)")

# df_pr = bảng làm việc trong bộ nhớ (mới nhất trước, khớp quy ước cũ) — không
# ghi đè price_history.csv, các cột tính toán bên dưới chỉ tồn tại trong biến
# này rồi xuất ra portfolio_history.csv / MyPortfolio.xlsx / data.json.
df_pr = price_hist.sort_values("date", ascending=False).reset_index(drop=True)


# ── 5. NẠP/RÚT TIỀN (W/D) — sổ cashflows.csv, thay cho input() thủ công ─────
# Trước đây bước này hỏi trực tiếp trên terminal. Vì chạy tự động trên GitHub
# Actions không có ai ngồi gõ, ta đọc từ manual_entries.csv thay thế.
# Khi cần thêm W/D: mở file manual_entries.csv trên GitHub (web), thêm dòng
# "date,type,value" (vd: 12/08/2026,D,10000000). Lần chạy tiếp theo sẽ tự gộp
# dòng đó vào data/cashflows.csv (sổ tích luỹ vĩnh viễn) rồi tự xoá dòng đã xử
# lý khỏi manual_entries.csv. cashflows.csv KHÔNG bị ghi đè mỗi lần chạy như
# portfolio_history.csv — nó chỉ được nối thêm (append), giống transactions.csv.
df_pr["date_str"] = df_pr["date"].dt.strftime("%d/%m/%Y")
valid_dates = set(df_pr["date_str"])


def load_cashflows(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["date", "type", "value"])
    df = pd.read_csv(path)
    return df if not df.empty else pd.DataFrame(columns=["date", "type", "value"])


def apply_manual_entries(cashflows_df: pd.DataFrame, manual_file: Path,
                          valid_dates: set) -> pd.DataFrame:
    if not manual_file.exists():
        return cashflows_df
    try:
        manual_df = pd.read_csv(manual_file)
    except pd.errors.EmptyDataError:
        return cashflows_df
    if manual_df.empty:
        return cashflows_df

    new_rows = []
    for _, row in manual_df.iterrows():
        date_str = str(row["date"]).strip()
        option   = str(row["type"]).strip().upper()
        if option not in ("W", "D"):
            print(f"⚠️  Bỏ qua dòng không hợp lệ (type='{option}')")
            continue
        try:
            value = float(row["value"])
        except (ValueError, TypeError):
            print(f"⚠️  Bỏ qua dòng không hợp lệ (value='{row['value']}')")
            continue
        if date_str not in valid_dates:
            print(f"⚠️  Không tìm thấy ngày {date_str} trong price_history.csv, bỏ qua")
            continue
        new_rows.append({"date": date_str, "type": option, "value": value})
        print(f"✅ Đã thêm {option} ngày {date_str} = {value:,.0f} vào cashflows.csv")

    if new_rows:
        cashflows_df = pd.concat([cashflows_df, pd.DataFrame(new_rows)], ignore_index=True)

    # Xoá các dòng đã xử lý, chỉ giữ lại header
    pd.DataFrame(columns=["date", "type", "value"]).to_csv(manual_file, index=False)
    return cashflows_df


cashflows = load_cashflows(CASHFLOWS_FILE)
cashflows = apply_manual_entries(cashflows, MANUAL_FILE, valid_dates)
CASHFLOWS_FILE.parent.mkdir(parents=True, exist_ok=True)
cashflows.to_csv(CASHFLOWS_FILE, index=False)

# Gộp W/D vào bảng làm việc. Nếu 1 ngày có nhiều dòng W (hoặc nhiều dòng D)
# trong cashflows.csv, chúng được CỘNG DỒN — khác với bản cũ (ghi đè cột, chỉ
# giữ được 1 giá trị/ngày/loại), đây là điểm sửa đúng hơn khi chuyển sang sổ
# ghi nhận dạng ledger.
if not cashflows.empty:
    wd = (
        cashflows.assign(value=cashflows["value"].astype(float))
        .groupby(["date", "type"])["value"].sum()
        .unstack(fill_value=0.0)
        .reindex(columns=["W", "D"], fill_value=0.0)
    )
else:
    wd = pd.DataFrame(columns=["W", "D"])

df_pr = df_pr.merge(wd, left_on="date_str", right_index=True, how="left")
df_pr[["W", "D"]] = df_pr[["W", "D"]].fillna(0.0)


# ── 6. TÍNH E1 (VECTORIZED) ───────────────────────────────────────────────────
# Cột nội bộ b,c,d,e,f,g,h tương ứng các mã cổ phiếu theo đúng thứ tự cộng vào
# E1 ở dưới (HPG, PNJ, TCB, MWG, MBB, FRT, FPT).
SYMBOL_TO_COL = {
    "HPG": "b", "PNJ": "c", "TCB": "d", "MWG": "e",
    "MBB": "f", "FRT": "g", "FPT": "h",
}
HOLDING_COLS = ["b", "c", "d", "e", "f", "g", "h", "margin"]


def load_transactions(path: Path) -> pd.DataFrame:
    """Đọc data/transactions.csv — đây là SOURCE OF TRUTH của danh mục.
    ⚠️ MỖI LẦN GIAO DỊCH: thêm 1 dòng mới vào file này (sửa trực tiếp trên
    GitHub web, không cần chạy code ở máy). Cột:
      date     : dd/mm/yyyy, ngày khớp lệnh
      symbol   : mã CK (HPG, PNJ, TCB, MWG, MBB, FRT, FPT) hoặc MARGIN
      action   : BUY / SELL (cho cổ phiếu)  hoặc  SET (cho MARGIN — đặt lại
                 số dư margin hiện tại, vì margin là số dư nợ chứ không phải
                 số lượng cộng dồn)
      quantity : số lượng CP (BUY/SELL) hoặc số dư margin mới (SET)
      note     : ghi chú tự do, không dùng để tính toán
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy {path}. Cần có data/transactions.csv làm nguồn "
            "dữ liệu giao dịch (xem cột date,symbol,action,quantity,note)."
        )
    tx = pd.read_csv(path, dtype={"note": str})
    if tx.empty:
        return tx
    tx["date"] = pd.to_datetime(tx["date"], format="%d/%m/%Y")
    unknown = ~tx["symbol"].isin(list(SYMBOL_TO_COL) + ["MARGIN"])
    if unknown.any():
        bad = tx.loc[unknown, "symbol"].unique().tolist()
        raise ValueError(f"transactions.csv có symbol không hợp lệ: {bad}")
    bad_action = ~(
        ((tx["symbol"] != "MARGIN") & tx["action"].isin(["BUY", "SELL"]))
        | ((tx["symbol"] == "MARGIN") & (tx["action"] == "SET"))
    )
    if bad_action.any():
        raise ValueError(
            "transactions.csv có action không hợp lệ ở các dòng:\n"
            f"{tx.loc[bad_action]}"
        )
    return tx.sort_values("date").reset_index(drop=True)


def get_holdings(dates: pd.Series, transactions: pd.DataFrame) -> pd.DataFrame:
    """Xây holdings (b,c,d,e,f,g,h,margin) theo từng ngày bằng cách cộng dồn
    (BUY/SELL) hoặc gán lại (SET, dùng cho margin) các giao dịch theo thời
    gian, rồi forward-fill sang các ngày sau đó — thay cho breakpoints
    hard-code trước đây."""
    result = pd.DataFrame(0.0, index=dates.index, columns=HOLDING_COLS)
    if transactions.empty:
        return result

    running = {c: 0.0 for c in HOLDING_COLS}
    for _, row in transactions.iterrows():
        if row["symbol"] == "MARGIN":
            col = "margin"
            running[col] = float(row["quantity"])
        else:
            col = SYMBOL_TO_COL[row["symbol"]]
            delta = row["quantity"] if row["action"] == "BUY" else -row["quantity"]
            running[col] += float(delta)
        result.loc[dates >= row["date"], col] = running[col]
    return result


dates_dt = pd.to_datetime(df_pr["date"])
transactions = load_transactions(TRANSACTIONS_FILE)
holdings = get_holdings(dates_dt, transactions)

df_pr["E1"] = (
      df_pr["HPG"] * holdings["b"]
    - holdings["margin"]
    + df_pr["PNJ"] * holdings["c"]
    + df_pr["TCB"] * holdings["d"]
    + df_pr["MWG"] * holdings["e"]
    + df_pr["MBB"] * holdings["f"]
    + df_pr["FRT"] * holdings["g"]
    + df_pr["FPT"] * holdings["h"]
)


# ── 6b. SỰ KIỆN GIAO DỊCH (BUY/SELL) CHO MARKER + TOOLTIP TRÊN CHART (mục 1) ─
# "Giá trị" của mỗi transaction KHÔNG tính qty×giá đóng cửa, mà lấy từ
# cashflows.csv theo ngày khớp (D↔BUY, W↔SELL) — vì danh mục không giữ tiền
# mặt nên D/W chính là số tiền dùng mua/bán CP hôm đó. Nếu 1 ngày có nhiều
# dòng D hoặc nhiều dòng W (đã cộng dồn trong biến `wd` ở mục 5), dùng tổng đó.
def build_transaction_events(transactions: pd.DataFrame, wd: pd.DataFrame) -> list:
    events = []
    tx_stock = transactions[transactions["symbol"] != "MARGIN"]
    for _, row in tx_stock.iterrows():
        date_str = row["date"].strftime("%d/%m/%Y")
        action   = row["action"]  # BUY hoặc SELL
        col      = "D" if action == "BUY" else "W"
        value    = float(wd.loc[date_str, col]) if (date_str in wd.index and col in wd.columns) else 0.0
        events.append({
            "date":     row["date"].strftime("%Y-%m-%d"),
            "symbol":   row["symbol"],
            "action":   action,
            "quantity": safe_num(row["quantity"]),
            "value":    safe_num(value),
        })
    return events


transaction_events = build_transaction_events(transactions, wd)


# ── 7. TÍNH CÁC CHỈ SỐ HIỆU SUẤT ─────────────────────────────────────────────
df_pr["E0"] = df_pr["E1"].shift(-1)
df_pr.loc[df_pr.index[-1], "E0"] = 0

df_pr["DR"]        = ((df_pr["E1"] + df_pr["W"]) - (df_pr["E0"] + df_pr["D"])) / (df_pr["E0"] + df_pr["D"])
df_pr["DR(VNI)"]   = df_pr["VNINDEX"].div(df_pr["VNINDEX"].shift(-1)).sub(1).fillna(0)
df_pr["DR+1"]      = df_pr["DR"] + 1
df_pr["DR+1(VNI)"] = df_pr["DR(VNI)"] + 1

df_pr["CR+1"]      = df_pr["DR+1"][::-1].cumprod()[::-1]
df_pr["CR+1(VNI)"] = df_pr["DR+1(VNI)"][::-1].cumprod()[::-1]
df_pr["CR"]        = df_pr["CR+1"] - 1
df_pr["CR(VNI)"]   = df_pr["CR+1(VNI)"] - 1


# ── 7b. TÍNH SHARPE RATIO & MAX DRAWDOWN DẠNG SỐ (mới, cho JSON dashboard) ──
df_pr["Sharpe"]      = np.nan
df_pr["MaxDrawdown"] = np.nan

# Cần tính năm (Year) trước để group
df_pr["Year_tmp"] = pd.to_datetime(df_pr["date"], format="%d/%m/%Y").dt.year
df_pr["YR"] = df_pr.groupby("Year_tmp")["DR+1"].transform(lambda s: s.cumprod().iloc[-1] - 1)

for _, idx in df_pr.groupby("YR").groups.items():
    idx = sorted(idx)
    dr_slice = df_pr.loc[idx, "DR"]
    std = dr_slice.std(ddof=0)
    if std:
        sharpe = (df_pr.loc[idx[0], "YR"] - RISK_FREE_RATE) / (std * np.sqrt(TRADING_DAYS))
        df_pr.loc[idx, "Sharpe"] = sharpe

    cr1_slice = pd.to_numeric(df_pr.loc[idx, "CR+1"], errors="coerce")[::-1]
    peaks = cr1_slice.cummax()
    max_dd = ((cr1_slice - peaks) / peaks).min()
    df_pr.loc[idx, "MaxDrawdown"] = max_dd


# ── 7c. SHARPE & MAX DRAWDOWN — BẢN TỔNG LỊCH SỬ (mục 3) ─────────────────────
# Không group theo năm — chạy trên toàn bộ chuỗi DR/CR+1. Sharpe tổng dùng CAGR
# (không dùng CR thô) vì CR trải nhiều năm sẽ méo tỷ lệ so với rf (lãi suất/năm).
N_total   = len(df_pr)
cr_latest = df_pr.loc[df_pr.index[0], "CR"]  # index 0 = ngày mới nhất (df_pr sort giảm dần)

if N_total and (1 + cr_latest) > 0:
    cagr_total = (1 + cr_latest) ** (TRADING_DAYS / N_total) - 1
else:
    cagr_total = np.nan  # (1+CR) <= 0 sẽ ra số phức với số mũ lẻ — coi như không xác định

std_total = df_pr["DR"].std(ddof=0)
sharpe_total = (
    (cagr_total - RISK_FREE_RATE) / (std_total * np.sqrt(TRADING_DAYS))
    if std_total and pd.notna(cagr_total) else np.nan
)

cr1_all   = pd.to_numeric(df_pr["CR+1"], errors="coerce")[::-1]  # đảo về thời gian tăng dần
peaks_all = cr1_all.cummax()
max_dd_total = ((cr1_all - peaks_all) / peaks_all).min()

df_pr["Sharpe_Total"]      = sharpe_total
df_pr["MaxDrawdown_Total"] = max_dd_total


# ── 7d. RISK/MARKET METRICS: BETA, ALPHA, VOLATILITY, CORRELATION (mục 2) ───
rf_daily = RISK_FREE_RATE / TRADING_DAYS


def _risk_metrics(dr: pd.Series, dr_vni: pd.Series) -> dict:
    """Beta/Alpha/Volatility/Correlation trên 1 lát cắt DR/DR(VNI). Alpha và
    Volatility luôn annualize bằng TRADING_DAYS (252) — kể cả khi lát cắt là
    bản Tổng nhiều năm hay 1 năm có số ngày lẻ (chốt thiết kế: không đổi 252)."""
    dr     = pd.to_numeric(dr, errors="coerce")
    dr_vni = pd.to_numeric(dr_vni, errors="coerce")
    dr_mean, vni_mean = dr.mean(), dr_vni.mean()

    cov     = ((dr - dr_mean) * (dr_vni - vni_mean)).mean()  # Cov quần thể (ddof=0)
    var_vni = ((dr_vni - vni_mean) ** 2).mean()               # Var quần thể, đồng bộ STDEVP
    beta    = cov / var_vni if var_vni else np.nan

    alpha_daily = dr_mean - (rf_daily + beta * (vni_mean - rf_daily)) if pd.notna(beta) else np.nan
    alpha       = alpha_daily * TRADING_DAYS if pd.notna(alpha_daily) else np.nan

    volatility  = dr.std(ddof=0) * np.sqrt(TRADING_DAYS)
    correlation = dr.corr(dr_vni)

    return {"beta": beta, "alpha": alpha, "volatility": volatility, "correlation": correlation}


df_pr["Beta"]        = np.nan
df_pr["Alpha"]       = np.nan
df_pr["Volatility"]  = np.nan
df_pr["Correlation"] = np.nan

for _, idx in df_pr.groupby("Year_tmp").groups.items():
    idx = sorted(idx)
    m = _risk_metrics(df_pr.loc[idx, "DR"], df_pr.loc[idx, "DR(VNI)"])
    df_pr.loc[idx, "Beta"]        = m["beta"]
    df_pr.loc[idx, "Alpha"]       = m["alpha"]
    df_pr.loc[idx, "Volatility"]  = m["volatility"]
    df_pr.loc[idx, "Correlation"] = m["correlation"]

risk_total = _risk_metrics(df_pr["DR"], df_pr["DR(VNI)"])
df_pr["Beta_Total"]        = risk_total["beta"]
df_pr["Alpha_Total"]       = risk_total["alpha"]
df_pr["Volatility_Total"]  = risk_total["volatility"]
df_pr["Correlation_Total"] = risk_total["correlation"]


# ── 7e. XIRR (mục 5) — lãi suất nội bộ trên chuỗi cashflow thật (D/W) ────────
# Quy ước dấu: D = âm (tiền ra mua CP), W = dương (tiền về từ bán CP) — khớp mục 1,
# vì danh mục không giữ tiền mặt nên D/W trong cashflows.csv chính là dòng tiền thật.
def _parsed_cashflows(cashflows_df: pd.DataFrame) -> list:
    parsed = []
    for _, row in cashflows_df.iterrows():
        try:
            d = datetime.strptime(str(row["date"]).strip(), "%d/%m/%Y").date()
        except ValueError:
            continue
        try:
            amt = float(row["value"])
        except (TypeError, ValueError):
            continue
        amt = -amt if str(row["type"]).strip().upper() == "D" else amt
        parsed.append((d, amt))
    return parsed


_cf_parsed = _parsed_cashflows(cashflows)

# XIRR tổng: toàn bộ D/W (cashflows.csv) + 1 dòng cuối = E1 hiện tại (dương), ngày mới nhất
_latest_idx  = df_pr.index[0]
_latest_date = df_pr.loc[_latest_idx, "date"].date()
_latest_e1   = float(df_pr.loc[_latest_idx, "E1"])

try:
    xirr_total = xirr(_cf_parsed + [(_latest_date, _latest_e1)])
except Exception:
    xirr_total = None

# XIRR theo năm: E0 đầu năm (âm, giá trị E1 tại ngày giao dịch đầu tiên của năm) +
# D/W phát sinh trong năm (nguyên dấu, không double-count với E0/E1 vì đó chỉ là
# mốc định giá) + E1 cuối năm (dương, giá trị E1 tại ngày giao dịch cuối năm).
df_pr["XIRR_Year"] = np.nan
for year, idx in df_pr.groupby("Year_tmp").groups.items():
    idx = sorted(idx)
    start_idx, end_idx = idx[-1], idx[0]  # df_pr giảm dần theo ngày: idx[-1]=sớm nhất, idx[0]=muộn nhất trong năm
    e0_date  = df_pr.loc[start_idx, "date"].date()
    e0_value = float(df_pr.loc[start_idx, "E1"])
    e1_date  = df_pr.loc[end_idx, "date"].date()
    e1_value = float(df_pr.loc[end_idx, "E1"])

    year_cfs = [(d, amt) for d, amt in _cf_parsed if d.year == year]
    cfs = [(e0_date, -e0_value)] + year_cfs + [(e1_date, e1_value)]

    try:
        xirr_year = xirr(cfs)
    except Exception:
        xirr_year = None
    df_pr.loc[idx, "XIRR_Year"] = xirr_year


# ── 7f. ALLOCATION + POSITION CONTRIBUTION (mục 4) — chỉ tính ngày mới nhất ──
def compute_positions(df_pr: pd.DataFrame, holdings: pd.DataFrame, transactions: pd.DataFrame) -> dict:
    if df_pr.empty:
        return {"as_of_date": None, "items": []}

    latest_idx  = df_pr.index[0]  # df_pr sắp xếp giảm dần theo ngày → dòng đầu = mới nhất
    latest_date = df_pr.loc[latest_idx, "date"]
    year        = latest_date.year

    year_rows = df_pr[df_pr["Year_tmp"] == year]
    year_start_idx = year_rows["date"].idxmin() if not year_rows.empty else None

    def price_on(symbol, dt):
        """Giá đóng cửa của `symbol` đúng ngày `dt` (khớp theo price_history)."""
        if dt is None or pd.isna(dt):
            return None
        rows = df_pr.loc[df_pr["date"] == dt, symbol]
        if rows.empty:
            return None
        v = rows.iloc[0]
        return float(v) if pd.notna(v) else None

    # Bước 1: qty/giá/giá trị từng mã đang có qty > 0 tại ngày mới nhất
    raw_positions = []
    total_value = 0.0
    for sym in DEFAULT_STOCK_CODES:
        col = SYMBOL_TO_COL[sym]
        qty = float(holdings.loc[latest_idx, col])
        if qty <= 0:
            continue
        price = float(df_pr.loc[latest_idx, sym])
        value = qty * price
        total_value += value
        raw_positions.append({"symbol": sym, "col": col, "qty": qty, "price": price, "value": value})

    # Bước 2: weight + return_year/return_total + contribution
    items = []
    for p in raw_positions:
        sym, col, qty, price, value = p["symbol"], p["col"], p["qty"], p["price"], p["value"]
        weight = value / total_value if total_value else None

        # return_year: giá mới nhất / giá ngày giao dịch đầu tiên của năm hiện tại - 1.
        # Nếu mã mới mua trong năm (chưa nắm giữ tại ngày đầu năm) thì dùng ngày BUY
        # đầu tiên của mã đó trong năm thay cho đầu năm.
        held_at_year_start = year_start_idx is not None and holdings.loc[year_start_idx, col] > 0
        if held_at_year_start:
            start_price_year = price_on(sym, df_pr.loc[year_start_idx, "date"])
        else:
            buys_this_year = transactions[
                (transactions["symbol"] == sym) & (transactions["action"] == "BUY") &
                (transactions["date"].dt.year == year)
            ]
            if not buys_this_year.empty:
                start_price_year = price_on(sym, buys_this_year["date"].min())
            elif year_start_idx is not None:
                start_price_year = price_on(sym, df_pr.loc[year_start_idx, "date"])
            else:
                start_price_year = None
        return_year = (price / start_price_year - 1) if start_price_year else None

        # return_total: giá mới nhất / giá tại ngày BUY đầu tiên (toàn bộ lịch sử, từ
        # transactions.csv, KHÔNG phải từ đầu price_history) - 1.
        buys_all = transactions[(transactions["symbol"] == sym) & (transactions["action"] == "BUY")]
        first_buy_total = buys_all["date"].min() if not buys_all.empty else None
        start_price_total = price_on(sym, first_buy_total) if first_buy_total is not None else None
        return_total = (price / start_price_total - 1) if start_price_total else None

        contribution_year  = weight * return_year  if (weight is not None and return_year  is not None) else None
        contribution_total = weight * return_total if (weight is not None and return_total is not None) else None

        items.append({
            "symbol": sym,
            "sector": SECTOR_MAP.get(sym),
            "qty":    safe_num(qty),
            "price":  safe_num(price),
            "value":  safe_num(value),
            "weight": safe_num(weight),
            "return_year":        safe_num(return_year),
            "return_total":       safe_num(return_total),
            "contribution_year":  safe_num(contribution_year),
            "contribution_total": safe_num(contribution_total),
        })

    return {"as_of_date": latest_date.strftime("%Y-%m-%d"), "items": items}


positions_data = compute_positions(df_pr, holdings, transactions)


# ── 8. XUẤT portfolio_history.csv (toàn bộ dữ liệu đã tính) ─────────────────
# File này là OUTPUT (derived), ghi đè hoàn toàn mỗi lần chạy — không phải nơi
# lưu trữ dữ liệu gốc nữa. Nguồn dữ liệu gốc thật sự là transactions.csv,
# price_history.csv và cashflows.csv (3 file đó mới cần backup/không được xoá).
df_pr["date"] = pd.to_datetime(df_pr["date"]).dt.strftime("%d/%m/%Y")
df_pr = df_pr.drop(columns=["date_str"], errors="ignore")

cols_to_drop = ["DR+1", "DR+1(VNI)", "CR+1", "CR+1(VNI)", "CR(VNI)",
                "MonthYear", "Cumulative_DR_M", "Year", "Year_tmp",
                "Cumulative_DR_Y", "Cumulative_DR_Y(VNI)"]
df_save = df_pr.drop(columns=[c for c in cols_to_drop if c in df_pr.columns])
PORTFOLIO_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
df_save.to_csv(PORTFOLIO_HISTORY_FILE, index=False)
print(f"✅ Đã lưu: {PORTFOLIO_HISTORY_FILE}")


# ── 9. THỐNG KÊ THÁNG / NĂM ──────────────────────────────────────────────────
df_pr["date"] = pd.to_datetime(df_pr["date"], format="%d/%m/%Y")

df_pr["MonthYear"] = df_pr["date"].dt.to_period("M")
df_pr["MR"] = df_pr.groupby("MonthYear")["DR+1"].transform(lambda s: s.cumprod().iloc[-1] - 1)

df_pr["Year"] = df_pr["date"].dt.year
df_pr["YR(VNI)"] = df_pr.groupby("Year")["DR+1(VNI)"].transform(lambda s: s.cumprod().iloc[-1] - 1)

df_pr["date"] = df_pr["date"].dt.strftime("%d/%m/%Y")


# ── 10. XUẤT MyPortfolio.xlsx ─────────────────────────────────────────────────
# Thứ tự cột cố định theo yêu cầu — "Sharpe ratio" và "Max drawdown" (dạng công
# thức Excel) được chèn thêm ở bước 11a ngay sau khi ghi file này, nên ở đây
# chỉ cần các cột số liệu gốc, không lấy 2 cột Sharpe/MaxDrawdown dạng số
# (chỉ dùng riêng cho JSON dashboard).
EXPORT_COLUMNS = [
    "date", "HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB", "VNINDEX",
    "E1", "W", "E0", "D", "DR", "DR(VNI)", "CR", "MR", "YR", "YR(VNI)",
]
df_export = df_pr[[c for c in EXPORT_COLUMNS if c in df_pr.columns]].copy()
df_export.to_excel(OUTPUT_FILE, index=False)


# ── 11. ĐỊNH DẠNG OPENPYXL ───────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb["Sheet1"]


def header_map(sheet) -> dict:
    return {str(c.value).strip(): c.column for c in sheet[1] if c.value}


hm = header_map(ws)


def merge_same_values(sheet, col_name: str, df_col: pd.Series, hm: dict):
    col_idx = hm.get(col_name)
    if col_idx is None:
        return
    letter = get_column_letter(col_idx)
    for val, idxs in df_col.groupby(df_col).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        sheet.merge_cells(f"{letter}{r1}:{letter}{r2}")


for col in ("MR", "YR", "YR(VNI)"):
    merge_same_values(ws, col, df_export[col] if col in df_export.columns else df_pr[col], hm)

yr_vni_idx = hm.get("YR(VNI)")
dr_idx     = hm.get("DR")
if yr_vni_idx and dr_idx:
    sharpe_idx    = yr_vni_idx + 1
    sharpe_letter = get_column_letter(sharpe_idx)
    dr_letter     = get_column_letter(dr_idx)
    ws[f"{sharpe_letter}1"] = "Sharpe ratio"

    df_yr = df_export["YR"] if "YR" in df_export.columns else df_pr["YR"]
    dr_col = df_export["DR"] if "DR" in df_export.columns else df_pr["DR"]
    for val, idxs in df_yr.groupby(df_yr).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        ws.merge_cells(f"{sharpe_letter}{r1}:{sharpe_letter}{r2}")
        ws[f"{sharpe_letter}{r1}"] = (
            f"=({val} - {RISK_FREE_RATE}) / "
            f"(STDEVP({dr_letter}{r1}:{dr_letter}{r2}) * SQRT({TRADING_DAYS}))"
        )

if yr_vni_idx:
    dd_idx    = yr_vni_idx + 2
    dd_letter = get_column_letter(dd_idx)
    ws[f"{dd_letter}1"] = "Max drawdown"

    df_cr1 = (df_export["CR+1"] if "CR+1" in df_export.columns else df_pr["CR+1"]).copy()
    df_yr2 = df_export["YR"] if "YR" in df_export.columns else df_pr["YR"]
    for val, idxs in df_yr2.groupby(df_yr2).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        ws.merge_cells(f"{dd_letter}{r1}:{dd_letter}{r2}")
        cr1_slice = pd.to_numeric(df_cr1.iloc[idxs[0]: idxs[-1] + 2], errors="coerce")[::-1]
        peaks     = cr1_slice.cummax()
        max_dd    = ((cr1_slice - peaks) / peaks).min()
        ws[f"{dd_letter}{r1}"] = max_dd


hm = header_map(ws)

date_style = NamedStyle(name="date_fmt", number_format="DD-MM-YYYY")
int_acc    = NamedStyle(name="int_acc",  number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
pct_style  = NamedStyle(name="pct_2",    number_format="0.00%")


def apply_style_to_cols(sheet, col_names, style, hm):
    for name in col_names:
        ci = hm.get(name)
        if ci:
            for row in sheet.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    cell.style = style


date_col = hm.get("date")
if date_col:
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        for cell in row:
            cell.style = date_style

apply_style_to_cols(ws, ["HPG", "TCB", "MWG", "MBB", "PNJ", "FRT", "FPT", "VNINDEX", "E1", "W", "E0", "D"], int_acc, hm)
apply_style_to_cols(ws, ["DR", "DR(VNI)", "CR", "MR", "YR", "YR(VNI)", "Max drawdown"], pct_style, hm)

sharpe_ci = hm.get("Sharpe ratio") or (yr_vni_idx + 1 if yr_vni_idx else None)
if sharpe_ci:
    for row in ws.iter_rows(min_row=1, min_col=sharpe_ci, max_col=sharpe_ci):
        for cell in row:
            cell.number_format = "0.00"

thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            cell.border = thin

COLOR_MAP = {
    "date": "C0FFC0", "HPG": "C0FFC0", "TCB": "C0FFC0", "FPT": "C0FFC0", "PNJ": "C0FFC0",
    "FRT": "CCFFFF", "MWG": "CCFFFF", "MBB": "CCFFFF",
    "VNINDEX": "FFA500",
    "E1": "4F81BD", "W": "4F81BD", "E0": "4F81BD", "D": "4F81BD",
    "DR": "8E7CC3", "DR(VNI)": "8E7CC3", "CR": "8E7CC3", "MR": "8E7CC3",
    "YR": "8E7CC3", "YR(VNI)": "8E7CC3", "Sharpe ratio": "8E7CC3", "Max drawdown": "8E7CC3",
}

hm = header_map(ws)
for col_name, hex_color in COLOR_MAP.items():
    ci = hm.get(col_name)
    if ci:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for row in ws.iter_rows(min_row=1, min_col=ci, max_col=ci):
            for cell in row:
                cell.fill = fill

src = ws["A1"]
for col in range(2, ws.max_column + 1):
    tgt = ws.cell(row=1, column=col)
    tgt.font      = copy(src.font)
    tgt.alignment = copy(src.alignment)

COLUMN_WIDTHS = {
    "date": 10.6, "HPG": 8.6, "TCB": 8.6, "MWG": 8.6, "FRT": 8.6,
    "FPT": 8.6, "MBB": 8.6, "PNJ": 8.6,
    "VNINDEX": 8.8, "E1": 11.0, "W": 11.0, "E0": 11.0, "D": 11.0,
    "DR": 8.1, "DR(VNI)": 8.1, "CR": 8.1, "MR": 8.1,
    "YR": 8.1, "YR(VNI)": 8.1, "Sharpe ratio": 11.2, "Max drawdown": 14.2,
}
hm = header_map(ws)
for name, width in COLUMN_WIDTHS.items():
    ci = hm.get(name)
    if ci:
        ws.column_dimensions[get_column_letter(ci)].width = width

wb.save(OUTPUT_FILE)
print(f"✅ Đã lưu: {OUTPUT_FILE}")


# ── 12. XUẤT JSON CHO WEB DASHBOARD ──────────────────────────────────────────
def export_json(df: pd.DataFrame, path: Path, events: list = None, positions: dict = None) -> None:
    records = df.copy()
    records["date"] = pd.to_datetime(records["date"], format="%d/%m/%Y")
    records = records.sort_values("date")
    records["date"] = records["date"].dt.strftime("%Y-%m-%d")

    keep_cols = [c for c in [
        "date", "HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB", "VNINDEX",
        "E1", "W", "E0", "D", "DR", "DR(VNI)", "CR", "CR(VNI)",
        "MR", "YR", "YR(VNI)", "Sharpe", "MaxDrawdown",
        "Sharpe_Total", "MaxDrawdown_Total",
        "Beta", "Alpha", "Volatility", "Correlation", "XIRR_Year",
    ] if c in records.columns]

    history = []
    for _, row in records[keep_cols].iterrows():
        rec = {"date": row["date"]}
        for c in keep_cols:
            if c == "date":
                continue
            rec[c] = safe_num(row[c])
        history.append(rec)

    latest = records.iloc[-1]
    summary = {
        "last_updated": datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "portfolio_value": safe_num(latest.get("E1")),
        "cumulative_return": safe_num(latest.get("CR")),
        "cumulative_return_vni": safe_num(latest.get("CR(VNI)")),
        "sharpe_ratio": safe_num(latest.get("Sharpe")),
        "max_drawdown": safe_num(latest.get("MaxDrawdown")),
        "sharpe_total": safe_num(latest.get("Sharpe_Total")),
        "max_drawdown_total": safe_num(latest.get("MaxDrawdown_Total")),
        "beta_total": safe_num(risk_total.get("beta")),
        "alpha_total": safe_num(risk_total.get("alpha")),
        "volatility_total": safe_num(risk_total.get("volatility")),
        "correlation_total": safe_num(risk_total.get("correlation")),
        "xirr_total": safe_num(xirr_total),
    }

    payload = {"summary": summary, "history": history}
    if events is not None:
        payload["events"] = events
    if positions is not None:
        payload["positions"] = positions

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, allow_nan=False)
    print(f"✅ Đã xuất: {path}")


export_json(df_pr, JSON_FILE, events=transaction_events, positions=positions_data)
