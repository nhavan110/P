# ── IMPORTS ──────────────────────────────────────────────────────────────────
import json
from copy import copy
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from vnstock import Vnstock

from dashboard_analytics import compute_dashboard_extras

# ── CẤU HÌNH ─────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).resolve().parent
PRICE_HISTORY_FILE     = BASE_DIR / "data" / "price_history.csv"      # giá thô (input/cache), KHÔNG chứa W/D hay kết quả
CASHFLOWS_FILE          = BASE_DIR / "data" / "cashflows.csv"          # sổ nạp/rút tiền (W/D), tích luỹ dần
PORTFOLIO_HISTORY_FILE  = BASE_DIR / "data" / "portfolio_history.csv"  # output: toàn bộ dữ liệu đã tính, ghi lại mỗi lần chạy
OUTPUT_FILE  = BASE_DIR / "data" / "MyPortfolio.xlsx"                  # output: bản Excel định dạng đẹp để xem/Drive
JSON_FILE    = BASE_DIR / "data" / "data.json"   # trong thư mục data/, khớp với fetch('data/data.json') trong index.html và được workflow "git add data/" commit
MANUAL_FILE  = BASE_DIR / "manual_entries.csv"       # inbox nhập W/D thủ công, tự xoá sau khi được gộp vào cashflows.csv
TRANSACTIONS_FILE = BASE_DIR / "data" / "transactions.csv"  # sổ giao dịch mua/bán CP + margin — SOURCE OF TRUTH của holdings
SHEET_NAME   = "Sheet1"
VNI_START    = "2022-06-01"
API_TEMPLATE = "https://api.simplize.vn/api/historical/quote/prices/{}?page=0&size=600"
DEFAULT_STOCK_CODES = ["HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB"]

RISK_FREE_RATE = 0.04
TRADING_DAYS   = 252


# ── 1. ĐỌC LỊCH SỬ GIÁ (price_history.csv — chỉ giá thô) ────────────────────
def load_price_history(path: Path) -> pd.DataFrame:
    cols = ["date", *DEFAULT_STOCK_CODES, "VNINDEX"]
    if not path.exists():
        return pd.DataFrame(columns=cols)
    df = pd.read_csv(path)
    df["date"] = pd.to_datetime(df["date"], format="%d/%m/%Y")
    return df.sort_values("date").reset_index(drop=True)


price_hist   = load_price_history(PRICE_HISTORY_FILE)
stock_codes  = [c for c in price_hist.columns if len(str(c)) == 3] or DEFAULT_STOCK_CODES


# ── 2. LẤY GIÁ CỔ PHIẾU TỪ API ───────────────────────────────────────────────
def fetch_stock_prices(codes: list) -> pd.DataFrame:
    """Lấy giá đóng cửa của danh sách mã CK, trả về DataFrame indexed by date."""
    frames = []
    for code in codes:
        resp = requests.get(API_TEMPLATE.format(code), timeout=10)
        resp.raise_for_status()
        df = pd.DataFrame(resp.json()["data"])
        df["date"] = pd.to_datetime(df["date"], unit="s")
        frames.append(df.set_index("date")["priceClose"].rename(code))
    return pd.concat(frames, axis=1).reset_index()   # outer join tự động


df_prices = fetch_stock_prices(stock_codes)


# ── 3. LẤY VNINDEX ────────────────────────────────────────────────────────────
today_str = date.today().strftime("%Y-%m-%d")
stock = Vnstock().stock(symbol="VNINDEX", source="VCI")
df_vni = (
    stock.quote.history(start=VNI_START, end=today_str)
    .rename(columns={"time": "date", "close": "VNINDEX"})
    .assign(date=lambda x: pd.to_datetime(x["date"]))
    .iloc[::-1]
    [["date", "VNINDEX"]]
)

df_prices["date"] = pd.to_datetime(df_prices["date"])
df_all = df_prices.merge(df_vni, on="date", how="left")


# ── 4. NỐI VỚI DỮ LIỆU CŨ, LƯU LẠI price_history.csv (chỉ giá thô) ──────────
last_update = price_hist["date"].max() if not price_hist.empty else (pd.Timestamp(VNI_START) - pd.Timedelta(days=1))

df_new     = df_all[df_all["date"] > last_update]
price_hist = (
    pd.concat([price_hist, df_new], ignore_index=True)
    .drop_duplicates(subset="date", keep="last")
    .sort_values("date")
    .fillna(0)
    .reset_index(drop=True)
)

price_hist_out = price_hist.copy()
price_hist_out["date"] = price_hist_out["date"].dt.strftime("%d/%m/%Y")
PRICE_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
price_hist_out.to_csv(PRICE_HISTORY_FILE, index=False)
print(f"✅ Đã lưu: {PRICE_HISTORY_FILE} ({len(price_hist_out)} dòng)")

# df_pr = bảng làm việc trong bộ nhớ (mới nhất trước, khớp quy ước cũ) — không
# ghi đè price_history.csv, các cột tính toán bên dưới chỉ tồn tại trong biến
# này rồi xuất ra portfolio_history.csv / MyPortfolio.xlsx / data.json.
df_pr = price_hist.sort_values("date", ascending=False).reset_index(drop=True)


# ── 5. NẠP/RÚT TIỀN (W/D) — sổ cashflows.csv, thay cho input() thủ công ─────
# Trước đây bước này hỏi trực tiếp trên terminal. Vì chạy tự động trên GitHub
# Actions không có ai ngồi gõ, ta đọc từ manual_entries.csv thay thế.
# Khi cần thêm W/D: mở file manual_entries.csv trên GitHub (web), thêm dòng
# "date,type,value" (vd: 12/08/2026,D,10000000). Lần chạy tiếp theo sẽ tự gộp
# dòng đó vào data/cashflows.csv (sổ tích luỹ vĩnh viễn) rồi tự xoá dòng đã xử
# lý khỏi manual_entries.csv. cashflows.csv KHÔNG bị ghi đè mỗi lần chạy như
# portfolio_history.csv — nó chỉ được nối thêm (append), giống transactions.csv.
df_pr["date_str"] = df_pr["date"].dt.strftime("%d/%m/%Y")
valid_dates = set(df_pr["date_str"])


def load_cashflows(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["date", "type", "value"])
    df = pd.read_csv(path)
    return df if not df.empty else pd.DataFrame(columns=["date", "type", "value"])


def apply_manual_entries(cashflows_df: pd.DataFrame, manual_file: Path,
                          valid_dates: set) -> pd.DataFrame:
    if not manual_file.exists():
        return cashflows_df
    try:
        manual_df = pd.read_csv(manual_file)
    except pd.errors.EmptyDataError:
        return cashflows_df
    if manual_df.empty:
        return cashflows_df

    new_rows = []
    for _, row in manual_df.iterrows():
        date_str = str(row["date"]).strip()
        option   = str(row["type"]).strip().upper()
        if option not in ("W", "D"):
            print(f"⚠️  Bỏ qua dòng không hợp lệ (type='{option}')")
            continue
        try:
            value = float(row["value"])
        except (ValueError, TypeError):
            print(f"⚠️  Bỏ qua dòng không hợp lệ (value='{row['value']}')")
            continue
        if date_str not in valid_dates:
            print(f"⚠️  Không tìm thấy ngày {date_str} trong price_history.csv, bỏ qua")
            continue
        new_rows.append({"date": date_str, "type": option, "value": value})
        print(f"✅ Đã thêm {option} ngày {date_str} = {value:,.0f} vào cashflows.csv")

    if new_rows:
        cashflows_df = pd.concat([cashflows_df, pd.DataFrame(new_rows)], ignore_index=True)

    # Xoá các dòng đã xử lý, chỉ giữ lại header
    pd.DataFrame(columns=["date", "type", "value"]).to_csv(manual_file, index=False)
    return cashflows_df


cashflows = load_cashflows(CASHFLOWS_FILE)
cashflows = apply_manual_entries(cashflows, MANUAL_FILE, valid_dates)
CASHFLOWS_FILE.parent.mkdir(parents=True, exist_ok=True)
cashflows.to_csv(CASHFLOWS_FILE, index=False)

# Gộp W/D vào bảng làm việc. Nếu 1 ngày có nhiều dòng W (hoặc nhiều dòng D)
# trong cashflows.csv, chúng được CỘNG DỒN — khác với bản cũ (ghi đè cột, chỉ
# giữ được 1 giá trị/ngày/loại), đây là điểm sửa đúng hơn khi chuyển sang sổ
# ghi nhận dạng ledger.
if not cashflows.empty:
    wd = (
        cashflows.assign(value=cashflows["value"].astype(float))
        .groupby(["date", "type"])["value"].sum()
        .unstack(fill_value=0.0)
        .reindex(columns=["W", "D"], fill_value=0.0)
    )
else:
    wd = pd.DataFrame(columns=["W", "D"])

df_pr = df_pr.merge(wd, left_on="date_str", right_index=True, how="left")
df_pr[["W", "D"]] = df_pr[["W", "D"]].fillna(0.0)


# ── 6. TÍNH E1 (VECTORIZED) ───────────────────────────────────────────────────
# Cột nội bộ b,c,d,e,f,g,h tương ứng các mã cổ phiếu theo đúng thứ tự cộng vào
# E1 ở dưới (HPG, PNJ, TCB, MWG, MBB, FRT, FPT).
SYMBOL_TO_COL = {
    "HPG": "b", "PNJ": "c", "TCB": "d", "MWG": "e",
    "MBB": "f", "FRT": "g", "FPT": "h",
}
HOLDING_COLS = ["b", "c", "d", "e", "f", "g", "h", "margin"]


def load_transactions(path: Path) -> pd.DataFrame:
    """Đọc data/transactions.csv — đây là SOURCE OF TRUTH của danh mục.
    ⚠️ MỖI LẦN GIAO DỊCH: thêm 1 dòng mới vào file này (sửa trực tiếp trên
    GitHub web, không cần chạy code ở máy). Cột:
      date     : dd/mm/yyyy, ngày khớp lệnh
      symbol   : mã CK (HPG, PNJ, TCB, MWG, MBB, FRT, FPT) hoặc MARGIN
      action   : BUY / SELL (cho cổ phiếu)  hoặc  SET (cho MARGIN — đặt lại
                 số dư margin hiện tại, vì margin là số dư nợ chứ không phải
                 số lượng cộng dồn)
      quantity : số lượng CP (BUY/SELL) hoặc số dư margin mới (SET)
      note     : ghi chú tự do, không dùng để tính toán
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy {path}. Cần có data/transactions.csv làm nguồn "
            "dữ liệu giao dịch (xem cột date,symbol,action,quantity,note)."
        )
    tx = pd.read_csv(path, dtype={"note": str})
    if tx.empty:
        return tx
    tx["date"] = pd.to_datetime(tx["date"], format="%d/%m/%Y")
    unknown = ~tx["symbol"].isin(list(SYMBOL_TO_COL) + ["MARGIN"])
    if unknown.any():
        bad = tx.loc[unknown, "symbol"].unique().tolist()
        raise ValueError(f"transactions.csv có symbol không hợp lệ: {bad}")
    bad_action = ~(
        ((tx["symbol"] != "MARGIN") & tx["action"].isin(["BUY", "SELL"]))
        | ((tx["symbol"] == "MARGIN") & (tx["action"] == "SET"))
    )
    if bad_action.any():
        raise ValueError(
            "transactions.csv có action không hợp lệ ở các dòng:\n"
            f"{tx.loc[bad_action]}"
        )
    return tx.sort_values("date").reset_index(drop=True)


def get_holdings(dates: pd.Series, transactions: pd.DataFrame) -> pd.DataFrame:
    """Xây holdings (b,c,d,e,f,g,h,margin) theo từng ngày bằng cách cộng dồn
    (BUY/SELL) hoặc gán lại (SET, dùng cho margin) các giao dịch theo thời
    gian, rồi forward-fill sang các ngày sau đó — thay cho breakpoints
    hard-code trước đây."""
    result = pd.DataFrame(0.0, index=dates.index, columns=HOLDING_COLS)
    if transactions.empty:
        return result

    running = {c: 0.0 for c in HOLDING_COLS}
    for _, row in transactions.iterrows():
        if row["symbol"] == "MARGIN":
            col = "margin"
            running[col] = float(row["quantity"])
        else:
            col = SYMBOL_TO_COL[row["symbol"]]
            delta = row["quantity"] if row["action"] == "BUY" else -row["quantity"]
            running[col] += float(delta)
        result.loc[dates >= row["date"], col] = running[col]
    return result


dates_dt = pd.to_datetime(df_pr["date"])
transactions = load_transactions(TRANSACTIONS_FILE)
holdings = get_holdings(dates_dt, transactions)

df_pr["E1"] = (
      df_pr["HPG"] * holdings["b"]
    - holdings["margin"]
    + df_pr["PNJ"] * holdings["c"]
    + df_pr["TCB"] * holdings["d"]
    + df_pr["MWG"] * holdings["e"]
    + df_pr["MBB"] * holdings["f"]
    + df_pr["FRT"] * holdings["g"]
    + df_pr["FPT"] * holdings["h"]
)


# ── 7. TÍNH CÁC CHỈ SỐ HIỆU SUẤT ─────────────────────────────────────────────
df_pr["E0"] = df_pr["E1"].shift(-1)
df_pr.loc[df_pr.index[-1], "E0"] = 0

df_pr["DR"]        = ((df_pr["E1"] + df_pr["W"]) - (df_pr["E0"] + df_pr["D"])) / (df_pr["E0"] + df_pr["D"])
df_pr["DR(VNI)"]   = df_pr["VNINDEX"].div(df_pr["VNINDEX"].shift(-1)).sub(1).fillna(0)
df_pr["DR+1"]      = df_pr["DR"] + 1
df_pr["DR+1(VNI)"] = df_pr["DR(VNI)"] + 1

df_pr["CR+1"]      = df_pr["DR+1"][::-1].cumprod()[::-1]
df_pr["CR+1(VNI)"] = df_pr["DR+1(VNI)"][::-1].cumprod()[::-1]
df_pr["CR"]        = df_pr["CR+1"] - 1
df_pr["CR(VNI)"]   = df_pr["CR+1(VNI)"] - 1


# ── 7b. TÍNH SHARPE RATIO & MAX DRAWDOWN DẠNG SỐ (mới, cho JSON dashboard) ──
df_pr["Sharpe"]      = np.nan
df_pr["MaxDrawdown"] = np.nan

# Cần tính năm (Year) trước để group
df_pr["Year_tmp"] = pd.to_datetime(df_pr["date"], format="%d/%m/%Y").dt.year
df_pr["YR"] = df_pr.groupby("Year_tmp")["DR+1"].transform(lambda s: s.cumprod().iloc[-1] - 1)

for _, idx in df_pr.groupby("YR").groups.items():
    idx = sorted(idx)
    dr_slice = df_pr.loc[idx, "DR"]
    std = dr_slice.std(ddof=0)
    if std:
        sharpe = (df_pr.loc[idx[0], "YR"] - RISK_FREE_RATE) / (std * np.sqrt(TRADING_DAYS))
        df_pr.loc[idx, "Sharpe"] = sharpe

    cr1_slice = pd.to_numeric(df_pr.loc[idx, "CR+1"], errors="coerce")[::-1]
    peaks = cr1_slice.cummax()
    max_dd = ((cr1_slice - peaks) / peaks).min()
    df_pr.loc[idx, "MaxDrawdown"] = max_dd


# ── 8. XUẤT portfolio_history.csv (toàn bộ dữ liệu đã tính) ─────────────────
# File này là OUTPUT (derived), ghi đè hoàn toàn mỗi lần chạy — không phải nơi
# lưu trữ dữ liệu gốc nữa. Nguồn dữ liệu gốc thật sự là transactions.csv,
# price_history.csv và cashflows.csv (3 file đó mới cần backup/không được xoá).
df_pr["date"] = pd.to_datetime(df_pr["date"]).dt.strftime("%d/%m/%Y")
df_pr = df_pr.drop(columns=["date_str"], errors="ignore")

cols_to_drop = ["DR+1", "DR+1(VNI)", "CR+1", "CR+1(VNI)", "CR(VNI)",
                "MonthYear", "Cumulative_DR_M", "Year", "Year_tmp",
                "Cumulative_DR_Y", "Cumulative_DR_Y(VNI)"]
df_save = df_pr.drop(columns=[c for c in cols_to_drop if c in df_pr.columns])
PORTFOLIO_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
df_save.to_csv(PORTFOLIO_HISTORY_FILE, index=False)
print(f"✅ Đã lưu: {PORTFOLIO_HISTORY_FILE}")


# ── 9. THỐNG KÊ THÁNG / NĂM ──────────────────────────────────────────────────
df_pr["date"] = pd.to_datetime(df_pr["date"], format="%d/%m/%Y")

df_pr["MonthYear"] = df_pr["date"].dt.to_period("M")
df_pr["MR"] = df_pr.groupby("MonthYear")["DR+1"].transform(lambda s: s.cumprod().iloc[-1] - 1)

df_pr["Year"] = df_pr["date"].dt.year
df_pr["YR(VNI)"] = df_pr.groupby("Year")["DR+1(VNI)"].transform(lambda s: s.cumprod().iloc[-1] - 1)

df_pr["date"] = df_pr["date"].dt.strftime("%d/%m/%Y")


# ── 10. XUẤT MyPortfolio.xlsx ─────────────────────────────────────────────────
# Thứ tự cột cố định theo yêu cầu — "Sharpe ratio" và "Max drawdown" (dạng công
# thức Excel) được chèn thêm ở bước 11a ngay sau khi ghi file này, nên ở đây
# chỉ cần các cột số liệu gốc, không lấy 2 cột Sharpe/MaxDrawdown dạng số
# (chỉ dùng riêng cho JSON dashboard).
EXPORT_COLUMNS = [
    "date", "HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB", "VNINDEX",
    "E1", "W", "E0", "D", "DR", "DR(VNI)", "CR", "MR", "YR", "YR(VNI)",
]
df_export = df_pr[[c for c in EXPORT_COLUMNS if c in df_pr.columns]].copy()
df_export.to_excel(OUTPUT_FILE, index=False)


# ── 11. ĐỊNH DẠNG OPENPYXL ───────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb["Sheet1"]


def header_map(sheet) -> dict:
    return {str(c.value).strip(): c.column for c in sheet[1] if c.value}


hm = header_map(ws)


def merge_same_values(sheet, col_name: str, df_col: pd.Series, hm: dict):
    col_idx = hm.get(col_name)
    if col_idx is None:
        return
    letter = get_column_letter(col_idx)
    for val, idxs in df_col.groupby(df_col).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        sheet.merge_cells(f"{letter}{r1}:{letter}{r2}")


for col in ("MR", "YR", "YR(VNI)"):
    merge_same_values(ws, col, df_export[col] if col in df_export.columns else df_pr[col], hm)

yr_vni_idx = hm.get("YR(VNI)")
dr_idx     = hm.get("DR")
if yr_vni_idx and dr_idx:
    sharpe_idx    = yr_vni_idx + 1
    sharpe_letter = get_column_letter(sharpe_idx)
    dr_letter     = get_column_letter(dr_idx)
    ws[f"{sharpe_letter}1"] = "Sharpe ratio"

    df_yr = df_export["YR"] if "YR" in df_export.columns else df_pr["YR"]
    dr_col = df_export["DR"] if "DR" in df_export.columns else df_pr["DR"]
    for val, idxs in df_yr.groupby(df_yr).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        ws.merge_cells(f"{sharpe_letter}{r1}:{sharpe_letter}{r2}")
        ws[f"{sharpe_letter}{r1}"] = (
            f"=({val} - {RISK_FREE_RATE}) / "
            f"(STDEVP({dr_letter}{r1}:{dr_letter}{r2}) * SQRT({TRADING_DAYS}))"
        )

if yr_vni_idx:
    dd_idx    = yr_vni_idx + 2
    dd_letter = get_column_letter(dd_idx)
    ws[f"{dd_letter}1"] = "Max drawdown"

    df_cr1 = (df_export["CR+1"] if "CR+1" in df_export.columns else df_pr["CR+1"]).copy()
    df_yr2 = df_export["YR"] if "YR" in df_export.columns else df_pr["YR"]
    for val, idxs in df_yr2.groupby(df_yr2).groups.items():
        idxs = sorted(idxs)
        r1, r2 = idxs[0] + 2, idxs[-1] + 2
        ws.merge_cells(f"{dd_letter}{r1}:{dd_letter}{r2}")
        cr1_slice = pd.to_numeric(df_cr1.iloc[idxs[0]: idxs[-1] + 2], errors="coerce")[::-1]
        peaks     = cr1_slice.cummax()
        max_dd    = ((cr1_slice - peaks) / peaks).min()
        ws[f"{dd_letter}{r1}"] = max_dd


hm = header_map(ws)

date_style = NamedStyle(name="date_fmt", number_format="DD-MM-YYYY")
int_acc    = NamedStyle(name="int_acc",  number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
pct_style  = NamedStyle(name="pct_2",    number_format="0.00%")


def apply_style_to_cols(sheet, col_names, style, hm):
    for name in col_names:
        ci = hm.get(name)
        if ci:
            for row in sheet.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    cell.style = style


date_col = hm.get("date")
if date_col:
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        for cell in row:
            cell.style = date_style

apply_style_to_cols(ws, ["HPG", "TCB", "MWG", "MBB", "PNJ", "FRT", "FPT", "VNINDEX", "E1", "W", "E0", "D"], int_acc, hm)
apply_style_to_cols(ws, ["DR", "DR(VNI)", "CR", "MR", "YR", "YR(VNI)", "Max drawdown"], pct_style, hm)

sharpe_ci = hm.get("Sharpe ratio") or (yr_vni_idx + 1 if yr_vni_idx else None)
if sharpe_ci:
    for row in ws.iter_rows(min_row=1, min_col=sharpe_ci, max_col=sharpe_ci):
        for cell in row:
            cell.number_format = "0.00"

thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            cell.border = thin

COLOR_MAP = {
    "date": "C0FFC0", "HPG": "C0FFC0", "TCB": "C0FFC0", "FPT": "C0FFC0", "PNJ": "C0FFC0",
    "FRT": "CCFFFF", "MWG": "CCFFFF", "MBB": "CCFFFF",
    "VNINDEX": "FFA500",
    "E1": "4F81BD", "W": "4F81BD", "E0": "4F81BD", "D": "4F81BD",
    "DR": "8E7CC3", "DR(VNI)": "8E7CC3", "CR": "8E7CC3", "MR": "8E7CC3",
    "YR": "8E7CC3", "YR(VNI)": "8E7CC3", "Sharpe ratio": "8E7CC3", "Max drawdown": "8E7CC3",
}

hm = header_map(ws)
for col_name, hex_color in COLOR_MAP.items():
    ci = hm.get(col_name)
    if ci:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for row in ws.iter_rows(min_row=1, min_col=ci, max_col=ci):
            for cell in row:
                cell.fill = fill

src = ws["A1"]
for col in range(2, ws.max_column + 1):
    tgt = ws.cell(row=1, column=col)
    tgt.font      = copy(src.font)
    tgt.alignment = copy(src.alignment)

COLUMN_WIDTHS = {
    "date": 10.6, "HPG": 8.6, "TCB": 8.6, "MWG": 8.6, "FRT": 8.6,
    "FPT": 8.6, "MBB": 8.6, "PNJ": 8.6,
    "VNINDEX": 8.8, "E1": 11.0, "W": 11.0, "E0": 11.0, "D": 11.0,
    "DR": 8.1, "DR(VNI)": 8.1, "CR": 8.1, "MR": 8.1,
    "YR": 8.1, "YR(VNI)": 8.1, "Sharpe ratio": 11.2, "Max drawdown": 14.2,
}
hm = header_map(ws)
for name, width in COLUMN_WIDTHS.items():
    ci = hm.get(name)
    if ci:
        ws.column_dimensions[get_column_letter(ci)].width = width

wb.save(OUTPUT_FILE)
print(f"✅ Đã lưu: {OUTPUT_FILE}")


# ── 12. XUẤT JSON CHO WEB DASHBOARD ──────────────────────────────────────────
def safe_num(x):
    """Trả về float hợp lệ, hoặc None nếu là NaN/Infinity (JSON chuẩn không
    chấp nhận NaN/Infinity — trình duyệt sẽ báo lỗi parse nếu để lọt vào)."""
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(v):
        return None
    return v


def export_json(df: pd.DataFrame, path: Path, transactions: pd.DataFrame) -> None:
    records = df.copy()
    records["date"] = pd.to_datetime(records["date"], format="%d/%m/%Y")
    records = records.sort_values("date")

    # ── Số liệu mở rộng cho 01/03/04/05/06/07 (positions, P&L, risk, drawdown, XIRR) ──
    extras = compute_dashboard_extras(records, transactions)

    records["date"] = records["date"].dt.strftime("%Y-%m-%d")

    keep_cols = [c for c in [
        "date", "HPG", "TCB", "FPT", "PNJ", "FRT", "MWG", "MBB", "VNINDEX",
        "E1", "W", "E0", "D", "DR", "DR(VNI)", "CR", "CR(VNI)",
        "MR", "YR", "YR(VNI)", "Sharpe", "MaxDrawdown",
    ] if c in records.columns]

    history = []
    for _, row in records[keep_cols].iterrows():
        rec = {"date": row["date"]}
        for c in keep_cols:
            if c == "date":
                continue
            rec[c] = safe_num(row[c])
        history.append(rec)

    latest = records.iloc[-1]
    summary = {
        "last_updated": datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "portfolio_value": safe_num(latest.get("E1")),
        "cumulative_return": safe_num(latest.get("CR")),
        "cumulative_return_vni": safe_num(latest.get("CR(VNI)")),
        "sharpe_ratio": safe_num(latest.get("Sharpe")),
        "max_drawdown": extras["max_drawdown_all"],
    }

    out = {
        "summary": summary,
        "history": history,
        "positions": extras["positions"],
        "sector_allocation": extras["sector_allocation"],
        "position_contribution": extras["position_contribution"],
        "transactions_markers": extras["transactions_markers"],
        "drawdown_series": extras["drawdown_series"],
        "current_drawdown": extras["current_drawdown"],
        "xirr_all": extras["xirr_all"],
        "periods": extras["periods"],
        "available_years": extras["available_years"],
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, allow_nan=False)
    print(f"✅ Đã xuất: {path}")


export_json(df_pr, JSON_FILE, transactions)
